import os
import io
import json
import uuid
from datetime import datetime
from math import ceil

from flask import Blueprint, render_template, redirect, url_for, request, flash, current_app, send_from_directory, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models import Supplier, RawMaterial, SemiFinished, FinishedProduct, Transaction
from app.decorators import admin_required
from app.utils.workflow_db import search_orders as wf_search_orders, get_order as wf_get_order, \
    get_order_stones, status_display as wf_status_display, design_type_display, stone_type_display

main_bp = Blueprint('main', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── 通用辅助 ──────────────────────────────────────────────

def _hide_cost_for_staff(items):
    """普通店员看不到成本字段"""
    if current_user.is_admin:
        return
    if isinstance(items, list):
        for item in items:
            if hasattr(item, 'cost_price'):
                item.cost_price = None
    elif hasattr(items, 'cost_price'):
        items.cost_price = None


def _save_photos(files, existing_photos=None):
    """保存上传的图片，返回 JSON 数组字符串"""
    photos = json.loads(existing_photos) if existing_photos else []
    upload_dir = current_app.config['UPLOAD_FOLDER']
    thumb_dir = current_app.config['THUMBNAIL_FOLDER']

    for f in files:
        if f and f.filename and allowed_file(f.filename):
            filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{f.filename}")
            filepath = os.path.join(upload_dir, filename)
            f.save(filepath)

            # 生成缩略图
            try:
                from PIL import Image
                img = Image.open(filepath)
                img.thumbnail(current_app.config['THUMBNAIL_SIZE'])
                thumb_path = os.path.join(thumb_dir, filename)
                img.save(thumb_path)
            except Exception:
                pass

            photos.append(filename)

    return json.dumps(photos, ensure_ascii=False)


# ── 上传文件服务 ──────────────────────────────────────────

@main_bp.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)


# ── 仪表盘 ────────────────────────────────────────────────

@main_bp.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('auth.login'))


@main_bp.route('/dashboard')
@login_required
def dashboard():
    raw_count = RawMaterial.query.count()
    semi_count = SemiFinished.query.count()
    finished_count = FinishedProduct.query.count()

    stats = {
        'raw_count': raw_count,
        'semi_count': semi_count,
        'finished_count': finished_count,
        'total': raw_count + semi_count + finished_count,
    }

    return render_template('dashboard.html', stats=stats)


# ══════════════════════════════════════════════════════════
# 供应商管理
# ══════════════════════════════════════════════════════════

@main_bp.route('/suppliers')
@login_required
def supplier_list():
    search = request.args.get('q', '').strip()
    type_filter = request.args.get('type', '').strip()
    query = Supplier.query

    if search:
        query = query.filter(
            db.or_(
                Supplier.name.contains(search),
                Supplier.contact_name.contains(search),
                Supplier.phone.contains(search),
            )
        )
    if type_filter:
        query = query.filter(Supplier.type == type_filter)

    suppliers = query.order_by(Supplier.name).all()
    return render_template('suppliers/list.html', suppliers=suppliers, search=search, type_filter=type_filter)


@main_bp.route('/suppliers/create', methods=['GET', 'POST'])
@login_required
def supplier_create():
    if request.method == 'POST':
        supplier = Supplier(
            name=request.form.get('name', '').strip(),
            type=request.form.get('type', 'stone'),
            contact_name=request.form.get('contact_name', '').strip() or None,
            phone=request.form.get('phone', '').strip() or None,
            wechat=request.form.get('wechat', '').strip() or None,
            address=request.form.get('address', '').strip() or None,
            notes=request.form.get('notes', '').strip() or None,
        )
        if not supplier.name:
            flash('供应商名称不能为空', 'warning')
            return render_template('suppliers/form.html', supplier=supplier)

        db.session.add(supplier)
        db.session.commit()
        flash(f'供应商「{supplier.name}」已添加', 'success')
        return redirect(url_for('main.supplier_list'))

    return render_template('suppliers/form.html', supplier=None)


@main_bp.route('/suppliers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def supplier_edit(id):
    supplier = Supplier.query.get_or_404(id)
    if request.method == 'POST':
        supplier.name = request.form.get('name', '').strip()
        supplier.type = request.form.get('type', 'stone')
        supplier.contact_name = request.form.get('contact_name', '').strip() or None
        supplier.phone = request.form.get('phone', '').strip() or None
        supplier.wechat = request.form.get('wechat', '').strip() or None
        supplier.address = request.form.get('address', '').strip() or None
        supplier.notes = request.form.get('notes', '').strip() or None
        supplier.updated_at = datetime.now()

        if not supplier.name:
            flash('供应商名称不能为空', 'warning')
            return render_template('suppliers/form.html', supplier=supplier)

        db.session.commit()
        flash(f'供应商「{supplier.name}」已更新', 'success')
        return redirect(url_for('main.supplier_list'))

    return render_template('suppliers/form.html', supplier=supplier)


@main_bp.route('/suppliers/<int:id>/toggle', methods=['POST'])
@login_required
def supplier_toggle(id):
    supplier = Supplier.query.get_or_404(id)
    supplier.is_active = not supplier.is_active
    supplier.updated_at = datetime.now()
    db.session.commit()
    status = '启用' if supplier.is_active else '停用'
    flash(f'供应商「{supplier.name}」已{status}', 'info')
    return redirect(url_for('main.supplier_list'))


# ══════════════════════════════════════════════════════════
# 原材料管理
# ══════════════════════════════════════════════════════════

@main_bp.route('/raw')
@login_required
def raw_list():
    search = request.args.get('q', '').strip()
    category = request.args.get('category', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 30

    query = RawMaterial.query

    if search:
        query = query.filter(
            db.or_(
                RawMaterial.code.contains(search),
                RawMaterial.name.contains(search),
                RawMaterial.cert_number.contains(search),
                RawMaterial.spec.contains(search),
                RawMaterial.location.contains(search),
                RawMaterial.lot_number.contains(search),
            )
        )
    if category:
        query = query.filter(RawMaterial.category == category)
    if status:
        query = query.filter(RawMaterial.status == status)

    query = query.order_by(RawMaterial.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items

    _hide_cost_for_staff(items)

    return render_template('raw/list.html',
                           items=items, pagination=pagination,
                           search=search, category=category, status=status)


@main_bp.route('/raw/create', methods=['GET', 'POST'])
@login_required
def raw_create():
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()

    if request.method == 'POST':
        photos_json = _save_photos(request.files.getlist('photos'))

        material = RawMaterial(
            code=request.form.get('code', '').strip() or None,
            category=request.form.get('category', 'stone'),
            name=request.form.get('name', '').strip(),
            spec=request.form.get('spec', '').strip() or None,

            weight=float(request.form.get('weight')) if request.form.get('weight') else None,
            unit=request.form.get('unit', 'ct'),

            shape=request.form.get('shape', '').strip() or None,
            color=request.form.get('color', '').strip() or None,
            clarity=request.form.get('clarity', '').strip() or None,
            cert_number=request.form.get('cert_number', '').strip() or None,
            cert_org=request.form.get('cert_org', '').strip() or None,

            purity=request.form.get('purity', '').strip() or None,

            cost_price=float(request.form.get('cost_price')) if request.form.get('cost_price') else None,
            purchase_date=datetime.strptime(request.form.get('purchase_date'), '%Y-%m-%d').date()
            if request.form.get('purchase_date') else None,

            supplier_id=int(request.form.get('supplier_id')) if request.form.get('supplier_id') else None,
            location=request.form.get('location', '').strip() or None,
            lot_number=request.form.get('lot_number', '').strip() or None,
            photos=photos_json if photos_json != '[]' else None,
            notes=request.form.get('notes', '').strip() or None,
            created_by=current_user.username,
        )

        if not material.name:
            flash('物料名称不能为空', 'warning')
            return render_template('raw/form.html', material=material, suppliers=suppliers)

        db.session.add(material)
        db.session.commit()
        flash(f'原材料「{material.name}」已录入', 'success')
        return redirect(url_for('main.raw_list'))

    return render_template('raw/form.html', material=None, suppliers=suppliers)


@main_bp.route('/raw/<int:id>')
@login_required
def raw_detail(id):
    material = RawMaterial.query.get_or_404(id)
    _hide_cost_for_staff(material)
    photos = json.loads(material.photos) if material.photos else []
    return render_template('raw/detail.html', material=material, photos=photos)


@main_bp.route('/raw/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def raw_edit(id):
    material = RawMaterial.query.get_or_404(id)
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()

    if request.method == 'POST':
        material.code = request.form.get('code', '').strip() or None
        material.category = request.form.get('category', 'stone')
        material.name = request.form.get('name', '').strip()
        material.spec = request.form.get('spec', '').strip() or None

        material.weight = float(request.form.get('weight')) if request.form.get('weight') else None
        material.unit = request.form.get('unit', 'ct')

        material.shape = request.form.get('shape', '').strip() or None
        material.color = request.form.get('color', '').strip() or None
        material.clarity = request.form.get('clarity', '').strip() or None
        material.cert_number = request.form.get('cert_number', '').strip() or None
        material.cert_org = request.form.get('cert_org', '').strip() or None

        material.purity = request.form.get('purity', '').strip() or None

        if request.form.get('cost_price'):
            material.cost_price = float(request.form.get('cost_price'))
        material.purchase_date = datetime.strptime(request.form.get('purchase_date'), '%Y-%m-%d').date() \
            if request.form.get('purchase_date') else None

        material.supplier_id = int(request.form.get('supplier_id')) if request.form.get('supplier_id') else None
        material.location = request.form.get('location', '').strip() or None
        material.status = request.form.get('status', 'in_stock')
        material.lot_number = request.form.get('lot_number', '').strip() or None
        material.notes = request.form.get('notes', '').strip() or None

        # 新照片追加到已有照片
        new_photos = _save_photos(request.files.getlist('photos'), material.photos)
        if new_photos != (material.photos or '[]'):
            material.photos = new_photos

        material.updated_at = datetime.now()

        if not material.name:
            flash('物料名称不能为空', 'warning')
            return render_template('raw/form.html', material=material, suppliers=suppliers)

        db.session.commit()
        flash(f'原材料「{material.name}」已更新', 'success')
        return redirect(url_for('main.raw_detail', id=material.id))

    return render_template('raw/form.html', material=material, suppliers=suppliers)


@main_bp.route('/raw/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def raw_delete(id):
    material = RawMaterial.query.get_or_404(id)
    name = material.name
    db.session.delete(material)
    db.session.commit()
    flash(f'原材料「{name}」已删除', 'info')
    return redirect(url_for('main.raw_list'))


@main_bp.route('/raw/<int:id>/copy')
@login_required
def raw_copy(id):
    """复制新建 — 除了编号清空，其他字段全部带入"""
    material = RawMaterial.query.get_or_404(id)
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()

    new_material = RawMaterial(
        code=None,
        category=material.category,
        name=material.name,
        spec=material.spec,
        weight=material.weight,
        unit=material.unit,
        shape=material.shape,
        color=material.color,
        clarity=material.clarity,
        cert_number=None,
        cert_org=material.cert_org,
        purity=material.purity,
        cost_price=material.cost_price if current_user.is_admin else None,
        supplier_id=material.supplier_id,
        location=material.location,
        lot_number=None,
        notes=material.notes,
    )
    flash(f'已复制「{material.name}」的信息，请修改后保存', 'info')
    return render_template('raw/form.html', material=new_material, suppliers=suppliers)


# ══════════════════════════════════════════════════════════
# 半成品管理
# ══════════════════════════════════════════════════════════

@main_bp.route('/semi')
@login_required
def semi_list():
    search = request.args.get('q', '').strip()
    type_filter = request.args.get('type', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 30

    query = SemiFinished.query

    if search:
        query = query.filter(
            db.or_(
                SemiFinished.code.contains(search),
                SemiFinished.name.contains(search),
                SemiFinished.craftsman.contains(search),
                SemiFinished.current_location.contains(search),
                SemiFinished.workflow_order_id.contains(search),
            )
        )
    if type_filter:
        query = query.filter(SemiFinished.type == type_filter)
    if status:
        query = query.filter(SemiFinished.status == status)

    query = query.order_by(SemiFinished.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items

    # 隐藏成本（管理员可见）
    for item in items:
        if not current_user.is_admin:
            item.cost_summary = None

    return render_template('semi/list.html',
                           items=items, pagination=pagination,
                           search=search, type_filter=type_filter, status=status)


@main_bp.route('/semi/create', methods=['GET', 'POST'])
@login_required
def semi_create():
    if request.method == 'POST':
        photos_json = _save_photos(request.files.getlist('photos'))

        item = SemiFinished(
            code=request.form.get('code', '').strip() or None,
            name=request.form.get('name', '').strip(),
            type=request.form.get('type', 'wax_model'),
            workflow_order_id=request.form.get('workflow_order_id', '').strip() or None,
            craftsman=request.form.get('craftsman', '').strip() or None,
            current_location=request.form.get('current_location', '').strip() or None,
            materials_snapshot=request.form.get('materials_snapshot', '').strip() or None,
            estimated_complete_date=datetime.strptime(
                request.form.get('estimated_complete_date'), '%Y-%m-%d').date()
            if request.form.get('estimated_complete_date') else None,
            cost_summary=float(request.form.get('cost_summary')) if request.form.get('cost_summary') else None,
            status=request.form.get('status', 'in_progress'),
            photos=photos_json if photos_json != '[]' else None,
            notes=request.form.get('notes', '').strip() or None,
            created_by=current_user.username,
        )

        if not item.name:
            flash('半成品名称不能为空', 'warning')
            return render_template('semi/form.html', item=item)

        db.session.add(item)
        db.session.commit()
        flash(f'半成品「{item.name}」已录入', 'success')
        return redirect(url_for('main.semi_list'))

    return render_template('semi/form.html', item=None)


@main_bp.route('/semi/<int:id>')
@login_required
def semi_detail(id):
    item = SemiFinished.query.get_or_404(id)
    if not current_user.is_admin:
        item.cost_summary = None
    photos = json.loads(item.photos) if item.photos else []
    return render_template('semi/detail.html', item=item, photos=photos)


@main_bp.route('/semi/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def semi_edit(id):
    item = SemiFinished.query.get_or_404(id)

    if request.method == 'POST':
        item.code = request.form.get('code', '').strip() or None
        item.name = request.form.get('name', '').strip()
        item.type = request.form.get('type', 'wax_model')
        item.workflow_order_id = request.form.get('workflow_order_id', '').strip() or None
        item.craftsman = request.form.get('craftsman', '').strip() or None
        item.current_location = request.form.get('current_location', '').strip() or None
        item.materials_snapshot = request.form.get('materials_snapshot', '').strip() or None
        item.estimated_complete_date = datetime.strptime(
            request.form.get('estimated_complete_date'), '%Y-%m-%d').date() \
            if request.form.get('estimated_complete_date') else None
        if request.form.get('cost_summary'):
            item.cost_summary = float(request.form.get('cost_summary'))
        item.status = request.form.get('status', 'in_progress')
        item.notes = request.form.get('notes', '').strip() or None

        new_photos = _save_photos(request.files.getlist('photos'), item.photos)
        if new_photos != (item.photos or '[]'):
            item.photos = new_photos

        item.updated_at = datetime.now()

        if not item.name:
            flash('半成品名称不能为空', 'warning')
            return render_template('semi/form.html', item=item)

        db.session.commit()
        flash(f'半成品「{item.name}」已更新', 'success')
        return redirect(url_for('main.semi_detail', id=item.id))

    return render_template('semi/form.html', item=item)


@main_bp.route('/semi/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def semi_delete(id):
    item = SemiFinished.query.get_or_404(id)
    name = item.name
    db.session.delete(item)
    db.session.commit()
    flash(f'半成品「{name}」已删除', 'info')
    return redirect(url_for('main.semi_list'))


@main_bp.route('/semi/<int:id>/copy')
@login_required
def semi_copy(id):
    item = SemiFinished.query.get_or_404(id)
    new_item = SemiFinished(
        code=None,
        name=item.name,
        type=item.type,
        workflow_order_id=None,
        craftsman=item.craftsman,
        current_location=item.current_location,
        materials_snapshot=item.materials_snapshot,
        cost_summary=item.cost_summary if current_user.is_admin else None,
        status='in_progress',
        notes=item.notes,
    )
    flash(f'已复制「{item.name}」的信息，请修改后保存', 'info')
    return render_template('semi/form.html', item=new_item)


# ══════════════════════════════════════════════════════════
# 成品管理
# ══════════════════════════════════════════════════════════

@main_bp.route('/finished')
@login_required
def finished_list():
    search = request.args.get('q', '').strip()
    type_filter = request.args.get('type', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 30

    query = FinishedProduct.query

    if search:
        query = query.filter(
            db.or_(
                FinishedProduct.product_code.contains(search),
                FinishedProduct.name.contains(search),
                FinishedProduct.main_stone.contains(search),
                FinishedProduct.location.contains(search),
                FinishedProduct.material_desc.contains(search),
            )
        )
    if type_filter:
        query = query.filter(FinishedProduct.type == type_filter)
    if status:
        query = query.filter(FinishedProduct.status == status)

    query = query.order_by(FinishedProduct.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items

    for item in items:
        if not current_user.is_admin:
            item.total_cost = None
            item.sale_price = None

    return render_template('finished/list.html',
                           items=items, pagination=pagination,
                           search=search, type_filter=type_filter, status=status)


@main_bp.route('/finished/create', methods=['GET', 'POST'])
@login_required
def finished_create():
    if request.method == 'POST':
        photos_json = _save_photos(request.files.getlist('photos'))

        product = FinishedProduct(
            product_code=request.form.get('product_code', '').strip() or None,
            name=request.form.get('name', '').strip(),
            type=request.form.get('type', '').strip() or None,
            material_desc=request.form.get('material_desc', '').strip() or None,

            gold_weight=float(request.form.get('gold_weight')) if request.form.get('gold_weight') else None,
            stone_weight=float(request.form.get('stone_weight')) if request.form.get('stone_weight') else None,

            main_stone=request.form.get('main_stone', '').strip() or None,
            side_stones=request.form.get('side_stones', '').strip() or None,

            total_cost=float(request.form.get('total_cost')) if request.form.get('total_cost') else None,
            sale_price=float(request.form.get('sale_price')) if request.form.get('sale_price') else None,

            location=request.form.get('location', '').strip() or None,
            status=request.form.get('status', 'in_stock'),
            workflow_order_id=request.form.get('workflow_order_id', '').strip() or None,
            photos=photos_json if photos_json != '[]' else None,
            notes=request.form.get('notes', '').strip() or None,
            created_by=current_user.username,
        )

        if not product.name:
            flash('成品名称不能为空', 'warning')
            return render_template('finished/form.html', product=product)

        db.session.add(product)
        db.session.commit()
        flash(f'成品「{product.name}」已录入', 'success')
        return redirect(url_for('main.finished_list'))

    return render_template('finished/form.html', product=None)


@main_bp.route('/finished/<int:id>')
@login_required
def finished_detail(id):
    product = FinishedProduct.query.get_or_404(id)
    if not current_user.is_admin:
        product.total_cost = None
        product.sale_price = None
    photos = json.loads(product.photos) if product.photos else []
    return render_template('finished/detail.html', product=product, photos=photos)


@main_bp.route('/finished/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def finished_edit(id):
    product = FinishedProduct.query.get_or_404(id)

    if request.method == 'POST':
        product.product_code = request.form.get('product_code', '').strip() or None
        product.name = request.form.get('name', '').strip()
        product.type = request.form.get('type', '').strip() or None
        product.material_desc = request.form.get('material_desc', '').strip() or None

        product.gold_weight = float(request.form.get('gold_weight')) if request.form.get('gold_weight') else None
        product.stone_weight = float(request.form.get('stone_weight')) if request.form.get('stone_weight') else None

        product.main_stone = request.form.get('main_stone', '').strip() or None
        product.side_stones = request.form.get('side_stones', '').strip() or None

        if request.form.get('total_cost'):
            product.total_cost = float(request.form.get('total_cost'))
        if request.form.get('sale_price'):
            product.sale_price = float(request.form.get('sale_price'))

        product.location = request.form.get('location', '').strip() or None
        product.status = request.form.get('status', 'in_stock')
        product.workflow_order_id = request.form.get('workflow_order_id', '').strip() or None
        product.notes = request.form.get('notes', '').strip() or None

        new_photos = _save_photos(request.files.getlist('photos'), product.photos)
        if new_photos != (product.photos or '[]'):
            product.photos = new_photos

        product.updated_at = datetime.now()

        if not product.name:
            flash('成品名称不能为空', 'warning')
            return render_template('finished/form.html', product=product)

        db.session.commit()
        flash(f'成品「{product.name}」已更新', 'success')
        return redirect(url_for('main.finished_detail', id=product.id))

    return render_template('finished/form.html', product=product)


@main_bp.route('/finished/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def finished_delete(id):
    product = FinishedProduct.query.get_or_404(id)
    name = product.name
    db.session.delete(product)
    db.session.commit()
    flash(f'成品「{name}」已删除', 'info')
    return redirect(url_for('main.finished_list'))


@main_bp.route('/finished/<int:id>/copy')
@login_required
def finished_copy(id):
    product = FinishedProduct.query.get_or_404(id)
    new_product = FinishedProduct(
        product_code=None,
        name=product.name,
        type=product.type,
        material_desc=product.material_desc,
        gold_weight=product.gold_weight,
        stone_weight=product.stone_weight,
        main_stone=product.main_stone,
        side_stones=product.side_stones,
        total_cost=product.total_cost if current_user.is_admin else None,
        sale_price=product.sale_price if current_user.is_admin else None,
        location=product.location,
        status='in_stock',
        workflow_order_id=None,
        notes=product.notes,
    )
    flash(f'已复制「{product.name}」的信息，请修改后保存', 'info')
    return render_template('finished/form.html', product=new_product)


# ══════════════════════════════════════════════════════════
# 全局搜索
# ══════════════════════════════════════════════════════════

@main_bp.route('/search')
@login_required
def global_search():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50

    if not q:
        return render_template('search.html', q='', results=[], total=0, pagination=None)

    # 搜索结果：三种类型各取
    raw_results = []
    semi_results = []
    finished_results = []

    # 原材料
    raw_query = RawMaterial.query.filter(
        db.or_(
            RawMaterial.code.contains(q),
            RawMaterial.name.contains(q),
            RawMaterial.cert_number.contains(q),
            RawMaterial.spec.contains(q),
            RawMaterial.location.contains(q),
            RawMaterial.lot_number.contains(q),
        )
    )
    for item in raw_query.limit(100).all():
        if not current_user.is_admin:
            item.cost_price = None
        raw_results.append({
            'type': 'raw', 'type_label': '原材料',
            'id': item.id, 'name': item.name, 'code': item.code,
            'desc': item.category_display + (' · ' + item.spec if item.spec else ''),
            'cert': item.cert_number, 'location': item.location,
            'status': item.status_display, 'status_class': item.status,
            'weight': item.weight_display,
        })

    # 半成品
    semi_query = SemiFinished.query.filter(
        db.or_(
            SemiFinished.code.contains(q),
            SemiFinished.name.contains(q),
            SemiFinished.craftsman.contains(q),
            SemiFinished.current_location.contains(q),
            SemiFinished.workflow_order_id.contains(q),
        )
    )
    for item in semi_query.limit(100).all():
        if not current_user.is_admin:
            item.cost_summary = None
        from_loc = item.current_location or ''
        if item.craftsman:
            from_loc = item.craftsman + (' · ' + from_loc if from_loc else '')
        semi_results.append({
            'type': 'semi', 'type_label': '半成品',
            'id': item.id, 'name': item.name, 'code': item.code,
            'desc': item.type_display,
            'cert': '', 'location': from_loc,
            'status': item.status_display, 'status_class': item.status,
            'weight': '',
        })

    # 成品
    finished_query = FinishedProduct.query.filter(
        db.or_(
            FinishedProduct.product_code.contains(q),
            FinishedProduct.name.contains(q),
            FinishedProduct.main_stone.contains(q),
            FinishedProduct.location.contains(q),
            FinishedProduct.material_desc.contains(q),
        )
    )
    for item in finished_query.limit(100).all():
        if not current_user.is_admin:
            item.total_cost = None
            item.sale_price = None
        weight = ''
        if item.gold_weight:
            weight += f'{item.gold_weight}g'
        if item.stone_weight:
            if weight:
                weight += ' / '
            weight += f'{item.stone_weight}ct'
        finished_results.append({
            'type': 'finished', 'type_label': '成品',
            'id': item.id, 'name': item.name, 'code': item.product_code,
            'desc': item.type_display + (' · ' + item.material_desc if item.material_desc else ''),
            'cert': item.main_stone or '', 'location': item.location,
            'status': item.status_display, 'status_class': item.status,
            'weight': weight,
        })

    results = raw_results + semi_results + finished_results
    total = len(results)

    # 简单分页
    start = (page - 1) * per_page
    end = start + per_page
    page_results = results[start:end]

    total_pages = ceil(total / per_page) if total > 0 else 1

    pagination_info = {
        'page': page, 'pages': total_pages, 'total': total,
        'has_prev': page > 1, 'has_next': page < total_pages,
        'prev_num': page - 1, 'next_num': page + 1,
    }

    return render_template('search.html', q=q, results=page_results,
                           total=total, pagination=pagination_info)


# ══════════════════════════════════════════════════════════
# Excel 批量导入 / 导出
# ══════════════════════════════════════════════════════════

# ── 模板列定义 ──────────────────────────────────────────

RAW_TEMPLATE = [
    ('code', '编号', 18), ('category', '分类', 12), ('name', '名称*', 20),
    ('spec', '规格描述', 20), ('weight', '重量', 10), ('unit', '单位(ct/g/颗)', 10),
    ('shape', '形状', 12), ('color', '颜色', 12), ('clarity', '净度', 12),
    ('cert_number', '证书号', 18), ('cert_org', '证书机构', 12),
    ('purity', '成色(金料)', 10), ('cost_price', '采购价', 12),
    ('purchase_date', '采购日期(YYYY-MM-DD)', 14),
    ('supplier_name', '供应商名称', 18), ('location', '存放位置', 18),
    ('lot_number', '批次号', 14), ('notes', '备注', 30),
]

SEMI_TEMPLATE = [
    ('code', '编号', 16), ('name', '名称*', 20),
    ('type', '类型', 12), ('workflow_order_id', '关联订单号', 16),
    ('craftsman', '当前工匠', 14), ('current_location', '当前所在', 18),
    ('materials_snapshot', '用料说明', 30),
    ('estimated_complete_date', '预计完成(YYYY-MM-DD)', 16),
    ('cost_summary', '成本汇总', 12), ('status', '状态', 12),
    ('notes', '备注', 30),
]

FINISHED_TEMPLATE = [
    ('product_code', '成品编号', 16), ('name', '名称*', 20),
    ('type', '类型', 12), ('material_desc', '材质描述', 18),
    ('gold_weight', '金重(g)', 10), ('stone_weight', '石重(ct)', 10),
    ('main_stone', '主石', 20), ('side_stones', '配石', 30),
    ('total_cost', '总成本(¥)', 12), ('sale_price', '售价(¥)', 12),
    ('location', '存放位置', 18), ('status', '状态', 10),
    ('workflow_order_id', '关联订单号', 16), ('notes', '备注', 30),
]


def _generate_template(columns, sheet_name='数据'):
    """生成 Excel 模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='D4A853', end_color='D4A853', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, (_, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    return wb


def _send_excel(wb, filename):
    """发送 Excel 文件流"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── 模板下载 ────────────────────────────────────────────

@main_bp.route('/raw/template')
@login_required
def raw_template():
    wb = _generate_template(RAW_TEMPLATE, '原材料导入模板')
    return _send_excel(wb, '原材料导入模板.xlsx')


@main_bp.route('/semi/template')
@login_required
def semi_template():
    wb = _generate_template(SEMI_TEMPLATE, '半成品导入模板')
    return _send_excel(wb, '半成品导入模板.xlsx')


@main_bp.route('/finished/template')
@login_required
def finished_template():
    wb = _generate_template(FINISHED_TEMPLATE, '成品导入模板')
    return _send_excel(wb, '成品导入模板.xlsx')


# ── 批量导入 ────────────────────────────────────────────

def _parse_import_file(file, columns, target_type):
    """解析上传的 Excel 文件，返回 (preview_rows, errors, raw_data)"""
    preview = []
    errors = []
    data = []

    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过标题行
        wb.close()
    except Exception as e:
        return [], [f'文件解析失败: {str(e)}'], []

    field_names = [c[0] for c in columns]

    for row_idx, row in enumerate(rows):
        if all(v is None or str(v).strip() == '' for v in row):
            continue  # 跳过空行

        row_data = {}
        row_display = {}
        row_errors = []

        for i, (field_name, _, _) in enumerate(columns):
            val = str(row[i]).strip() if i < len(row) and row[i] is not None else ''
            row_data[field_name] = val if val else None
            row_display[field_name] = val if val else ''

        # 校验必填字段
        if target_type == 'raw' and not row_data.get('name'):
            row_errors.append('名称不能为空')
        elif target_type == 'semi' and not row_data.get('name'):
            row_errors.append('名称不能为空')
        elif target_type == 'finished' and not row_data.get('name'):
            row_errors.append('名称不能为空')

        # 校验数值字段
        numeric_fields = []
        if target_type == 'raw':
            numeric_fields = ['weight', 'cost_price']
        elif target_type == 'semi':
            numeric_fields = ['cost_summary']
        elif target_type == 'finished':
            numeric_fields = ['gold_weight', 'stone_weight', 'total_cost', 'sale_price']

        for nf in numeric_fields:
            if row_data.get(nf):
                try:
                    float(row_data[nf])
                except ValueError:
                    row_errors.append(f'{nf}: 请输入有效数字')

        # 校验日期
        date_fields = []
        if target_type in ('raw',):
            date_fields = ['purchase_date']
        elif target_type in ('semi',):
            date_fields = ['estimated_complete_date']

        for df in date_fields:
            if row_data.get(df):
                try:
                    datetime.strptime(row_data[df], '%Y-%m-%d')
                except ValueError:
                    row_errors.append(f'{df}: 日期格式应为 YYYY-MM-DD')

        preview.append({
            'row_num': row_idx + 2,
            'data': row_display,
            'errors': row_errors,
            'ok': len(row_errors) == 0,
        })
        if not row_errors:
            data.append(row_data)

    return preview, errors, data


# 导入缓存（简单实现，存 session 太大，用临时文件）
_import_cache = {}  # {cache_key: [data_list]}


@main_bp.route('/raw/import', methods=['GET', 'POST'])
@login_required
def raw_import():
    """原材料批量导入"""
    preview_rows = None
    error_count = 0
    ok_count = 0
    cache_key = None

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('请选择要上传的 Excel 文件', 'warning')
        elif not file.filename.endswith(('.xlsx', '.xls')):
            flash('仅支持 .xlsx 文件', 'warning')
        else:
            preview_rows, errors, data = _parse_import_file(file, RAW_TEMPLATE, 'raw')
            if errors:
                flash(errors[0], 'danger')
                return redirect(url_for('main.raw_import'))

            ok_count = sum(1 for r in preview_rows if r['ok'])
            error_count = sum(1 for r in preview_rows if not r['ok'])

            if ok_count > 0:
                cache_key = str(uuid.uuid4())
                _import_cache[cache_key] = data

    return render_template('import_form.html',
                           target_type='raw', target_label='原材料',
                           preview_rows=preview_rows, ok_count=ok_count,
                           error_count=error_count, cache_key=cache_key,
                           import_url=url_for('main.raw_import_confirm'),
                           template_url=url_for('main.raw_template'),
                           list_url=url_for('main.raw_list'))


@main_bp.route('/raw/import/confirm', methods=['POST'])
@login_required
def raw_import_confirm():
    """确认导入原材料"""
    cache_key = request.form.get('cache_key', '')
    data = _import_cache.pop(cache_key, [])
    if not data:
        flash('导入数据已过期，请重新上传', 'warning')
        return redirect(url_for('main.raw_import'))

    count = 0
    for row in data:
        material = RawMaterial(
            code=row.get('code'), category=row.get('category', 'stone'),
            name=row['name'], spec=row.get('spec'),
            weight=float(row['weight']) if row.get('weight') else None,
            unit=row.get('unit', 'ct'),
            shape=row.get('shape'), color=row.get('color'),
            clarity=row.get('clarity'),
            cert_number=row.get('cert_number'), cert_org=row.get('cert_org'),
            purity=row.get('purity'),
            cost_price=float(row['cost_price']) if row.get('cost_price') else None,
            purchase_date=datetime.strptime(row['purchase_date'], '%Y-%m-%d').date()
            if row.get('purchase_date') else None,
            location=row.get('location'), lot_number=row.get('lot_number'),
            notes=row.get('notes'), created_by=current_user.username,
        )
        # 处理供应商名称 → ID
        supplier_name = row.get('supplier_name')
        if supplier_name:
            supplier = Supplier.query.filter_by(name=supplier_name).first()
            if supplier:
                material.supplier_id = supplier.id

        db.session.add(material)
        count += 1

    db.session.commit()
    flash(f'成功导入 {count} 条原材料', 'success')
    return redirect(url_for('main.raw_list'))


@main_bp.route('/semi/import', methods=['GET', 'POST'])
@login_required
def semi_import():
    preview_rows = None
    error_count = 0
    ok_count = 0
    cache_key = None

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('请选择要上传的 Excel 文件', 'warning')
        elif not file.filename.endswith(('.xlsx', '.xls')):
            flash('仅支持 .xlsx 文件', 'warning')
        else:
            preview_rows, errors, data = _parse_import_file(file, SEMI_TEMPLATE, 'semi')
            if errors:
                flash(errors[0], 'danger')
                return redirect(url_for('main.semi_import'))
            ok_count = sum(1 for r in preview_rows if r['ok'])
            error_count = sum(1 for r in preview_rows if not r['ok'])
            if ok_count > 0:
                cache_key = str(uuid.uuid4())
                _import_cache[cache_key] = data

    return render_template('import_form.html',
                           target_type='semi', target_label='半成品',
                           preview_rows=preview_rows, ok_count=ok_count,
                           error_count=error_count, cache_key=cache_key,
                           import_url=url_for('main.semi_import_confirm'),
                           template_url=url_for('main.semi_template'),
                           list_url=url_for('main.semi_list'))


@main_bp.route('/semi/import/confirm', methods=['POST'])
@login_required
def semi_import_confirm():
    cache_key = request.form.get('cache_key', '')
    data = _import_cache.pop(cache_key, [])
    if not data:
        flash('导入数据已过期，请重新上传', 'warning')
        return redirect(url_for('main.semi_import'))

    count = 0
    for row in data:
        item = SemiFinished(
            code=row.get('code'), name=row['name'],
            type=row.get('type', 'wax_model'),
            workflow_order_id=row.get('workflow_order_id'),
            craftsman=row.get('craftsman'),
            current_location=row.get('current_location'),
            materials_snapshot=row.get('materials_snapshot'),
            estimated_complete_date=datetime.strptime(
                row['estimated_complete_date'], '%Y-%m-%d').date()
            if row.get('estimated_complete_date') else None,
            cost_summary=float(row['cost_summary']) if row.get('cost_summary') else None,
            status=row.get('status', 'in_progress'),
            notes=row.get('notes'), created_by=current_user.username,
        )
        db.session.add(item)
        count += 1

    db.session.commit()
    flash(f'成功导入 {count} 条半成品', 'success')
    return redirect(url_for('main.semi_list'))


@main_bp.route('/finished/import', methods=['GET', 'POST'])
@login_required
def finished_import():
    preview_rows = None
    error_count = 0
    ok_count = 0
    cache_key = None

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('请选择要上传的 Excel 文件', 'warning')
        elif not file.filename.endswith(('.xlsx', '.xls')):
            flash('仅支持 .xlsx 文件', 'warning')
        else:
            preview_rows, errors, data = _parse_import_file(file, FINISHED_TEMPLATE, 'finished')
            if errors:
                flash(errors[0], 'danger')
                return redirect(url_for('main.finished_import'))
            ok_count = sum(1 for r in preview_rows if r['ok'])
            error_count = sum(1 for r in preview_rows if not r['ok'])
            if ok_count > 0:
                cache_key = str(uuid.uuid4())
                _import_cache[cache_key] = data

    return render_template('import_form.html',
                           target_type='finished', target_label='成品',
                           preview_rows=preview_rows, ok_count=ok_count,
                           error_count=error_count, cache_key=cache_key,
                           import_url=url_for('main.finished_import_confirm'),
                           template_url=url_for('main.finished_template'),
                           list_url=url_for('main.finished_list'))


@main_bp.route('/finished/import/confirm', methods=['POST'])
@login_required
def finished_import_confirm():
    cache_key = request.form.get('cache_key', '')
    data = _import_cache.pop(cache_key, [])
    if not data:
        flash('导入数据已过期，请重新上传', 'warning')
        return redirect(url_for('main.finished_import'))

    count = 0
    for row in data:
        product = FinishedProduct(
            product_code=row.get('product_code'), name=row['name'],
            type=row.get('type'), material_desc=row.get('material_desc'),
            gold_weight=float(row['gold_weight']) if row.get('gold_weight') else None,
            stone_weight=float(row['stone_weight']) if row.get('stone_weight') else None,
            main_stone=row.get('main_stone'), side_stones=row.get('side_stones'),
            total_cost=float(row['total_cost']) if row.get('total_cost') else None,
            sale_price=float(row['sale_price']) if row.get('sale_price') else None,
            location=row.get('location'), status=row.get('status', 'in_stock'),
            workflow_order_id=row.get('workflow_order_id'),
            notes=row.get('notes'), created_by=current_user.username,
        )
        db.session.add(product)
        count += 1

    db.session.commit()
    flash(f'成功导入 {count} 条成品', 'success')
    return redirect(url_for('main.finished_list'))


# ── Excel 导出 ────────────────────────────────────────────

def _export_to_excel(columns, items, row_mapper, filename):
    """通用导出"""
    wb = Workbook()
    ws = wb.active
    ws.title = '导出数据'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='D4A853', end_color='D4A853', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, (_, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, item in enumerate(items, 2):
        values = row_mapper(item)
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.freeze_panes = 'A2'
    return _send_excel(wb, filename)


@main_bp.route('/raw/export')
@login_required
def raw_export():
    search = request.args.get('q', '').strip()
    category = request.args.get('category', '').strip()
    status = request.args.get('status', '').strip()

    query = RawMaterial.query
    if search:
        query = query.filter(db.or_(
            RawMaterial.code.contains(search),
            RawMaterial.name.contains(search),
            RawMaterial.cert_number.contains(search),
            RawMaterial.spec.contains(search),
        ))
    if category:
        query = query.filter(RawMaterial.category == category)
    if status:
        query = query.filter(RawMaterial.status == status)
    items = query.order_by(RawMaterial.created_at.desc()).all()

    def mapper(item):
        return [
            item.code, item.name, item.category_display, item.spec,
            item.weight, item.unit, item.shape, item.color, item.clarity,
            item.cert_number, item.cert_org, item.purity,
            item.cost_price if current_user.is_admin else '',
            str(item.purchase_date) if item.purchase_date else '',
            item.supplier.name if item.supplier else '',
            item.location, item.lot_number, item.status_display, item.notes,
        ]

    export_columns = RAW_TEMPLATE + [('status', '状态', 12)]
    return _export_to_excel(export_columns, items, mapper, '原材料库存导出.xlsx')


@main_bp.route('/semi/export')
@login_required
def semi_export():
    search = request.args.get('q', '').strip()
    type_filter = request.args.get('type', '').strip()
    status = request.args.get('status', '').strip()

    query = SemiFinished.query
    if search:
        query = query.filter(db.or_(
            SemiFinished.code.contains(search),
            SemiFinished.name.contains(search),
            SemiFinished.craftsman.contains(search),
        ))
    if type_filter:
        query = query.filter(SemiFinished.type == type_filter)
    if status:
        query = query.filter(SemiFinished.status == status)
    items = query.order_by(SemiFinished.created_at.desc()).all()

    def mapper(item):
        return [
            item.code, item.name, item.type_display,
            item.workflow_order_id, item.craftsman,
            item.current_location, item.materials_snapshot,
            str(item.estimated_complete_date) if item.estimated_complete_date else '',
            item.cost_summary if current_user.is_admin else '',
            item.status_display, item.notes,
        ]

    export_columns = SEMI_TEMPLATE
    return _export_to_excel(export_columns, items, mapper, '半成品库存导出.xlsx')


@main_bp.route('/finished/export')
@login_required
def finished_export():
    search = request.args.get('q', '').strip()
    type_filter = request.args.get('type', '').strip()
    status = request.args.get('status', '').strip()

    query = FinishedProduct.query
    if search:
        query = query.filter(db.or_(
            FinishedProduct.product_code.contains(search),
            FinishedProduct.name.contains(search),
            FinishedProduct.main_stone.contains(search),
        ))
    if type_filter:
        query = query.filter(FinishedProduct.type == type_filter)
    if status:
        query = query.filter(FinishedProduct.status == status)
    items = query.order_by(FinishedProduct.created_at.desc()).all()

    def mapper(item):
        return [
            item.product_code, item.name, item.type_display,
            item.material_desc, item.gold_weight, item.stone_weight,
            item.main_stone, item.side_stones,
            item.total_cost if current_user.is_admin else '',
            item.sale_price if current_user.is_admin else '',
            item.location, item.status_display,
            item.workflow_order_id, item.notes,
        ]

    export_columns = FINISHED_TEMPLATE
    return _export_to_excel(export_columns, items, mapper, '成品库存导出.xlsx')


# ══════════════════════════════════════════════════════════
# 出入库流水
# ══════════════════════════════════════════════════════════

def _get_item(target_type, target_id):
    """根据类型和ID获取物料对象"""
    model_map = {'raw': RawMaterial, 'semi': SemiFinished, 'finished': FinishedProduct}
    model = model_map.get(target_type)
    if not model:
        return None
    return model.query.get(target_id)


def _update_item_status(target_type, target_id, trans_type, to_location=None, related_order_id=None):
    """根据交易类型更新物料状态"""
    item = _get_item(target_type, target_id)
    if not item:
        return

    if target_type == 'raw':
        status_field = 'status'
        location_field = 'location'
    elif target_type == 'semi':
        status_field = 'status'
        location_field = 'current_location'
    else:
        status_field = 'status'
        location_field = 'location'

    if trans_type == 'in':
        setattr(item, status_field, 'in_stock')
        if to_location:
            setattr(item, location_field, to_location)
    elif trans_type == 'out':
        setattr(item, status_field, 'out')
    elif trans_type == 'transfer':
        if to_location:
            setattr(item, location_field, to_location)
    elif trans_type == 'lock':
        setattr(item, status_field, 'locked')
        if target_type == 'semi' and related_order_id:
            setattr(item, 'workflow_order_id', related_order_id)
        elif target_type == 'finished' and related_order_id:
            setattr(item, 'workflow_order_id', related_order_id)
    elif trans_type == 'unlock':
        setattr(item, status_field, 'in_stock')

    item.updated_at = datetime.now()


@main_bp.route('/transactions')
@login_required
def transaction_list():
    trans_type = request.args.get('type', '').strip()
    target_type = request.args.get('target_type', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50

    query = Transaction.query

    if trans_type:
        query = query.filter(Transaction.type == trans_type)
    if target_type:
        query = query.filter(Transaction.target_type == target_type)

    query = query.order_by(Transaction.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items

    # 普通店员隐藏成本
    if not current_user.is_admin:
        for item in items:
            item.cost_recorded = None

    return render_template('transactions/list.html',
                           items=items, pagination=pagination,
                           type_filter=trans_type, target_type_filter=target_type)


@main_bp.route('/transactions/create', methods=['GET', 'POST'])
@login_required
def transaction_create():
    # 预填参数
    prefilled = {
        'target_type': request.args.get('target_type', 'raw'),
        'target_id': request.args.get('target_id', ''),
        'action': request.args.get('action', 'in'),
    }

    if request.method == 'POST':
        trans_type = request.form.get('type', 'in')
        target_type = request.form.get('target_type', 'raw')
        target_id = int(request.form.get('target_id', 0))

        item = _get_item(target_type, target_id)
        if not item:
            flash('未找到对应物料', 'danger')
            return redirect(url_for('main.transaction_list'))

        # 获取物料名称
        target_name = item.name
        weight = getattr(item, 'weight', None) or getattr(item, 'gold_weight', None)

        trans = Transaction(
            type=trans_type,
            target_type=target_type,
            target_id=target_id,
            target_name=target_name,
            quantity=1,
            weight=float(request.form.get('weight')) if request.form.get('weight') else weight,
            from_location=getattr(item, 'location', None) or getattr(item, 'current_location', None),
            to_location=request.form.get('to_location', '').strip() or None,
            related_order_id=request.form.get('related_order_id', '').strip() or None,
            cost_recorded=float(request.form.get('cost_recorded')) if request.form.get('cost_recorded') else None,
            operator=current_user.username,
            reason=request.form.get('reason', '').strip() or None,
        )

        db.session.add(trans)
        db.session.flush()

        # 更新物料状态
        _update_item_status(
            target_type, target_id, trans_type,
            to_location=trans.to_location,
            related_order_id=trans.related_order_id,
        )

        db.session.commit()
        flash(f'{trans.type_display}记录已保存', 'success')
        return redirect(url_for('main.transaction_list'))

    return render_template('transactions/form.html', prefilled=prefilled)


@main_bp.route('/transactions/<int:id>')
@login_required
def transaction_detail(id):
    trans = Transaction.query.get_or_404(id)
    if not current_user.is_admin:
        trans.cost_recorded = None
    return render_template('transactions/detail.html', trans=trans)


# ══════════════════════════════════════════════════════════
# 快捷操作（从详情页直接操作）
# ══════════════════════════════════════════════════════════

def _quick_action(target_type, target_id, action):
    """通用快捷操作: 入库/出库/锁定/解锁"""
    item = _get_item(target_type, target_id)
    if not item:
        flash('物料不存在', 'danger')
        return None

    item_name = item.name
    location = getattr(item, 'location', None) or getattr(item, 'current_location', None)
    weight = getattr(item, 'weight', None) or getattr(item, 'gold_weight', None)

    trans = Transaction(
        type=action,
        target_type=target_type,
        target_id=target_id,
        target_name=item_name,
        quantity=1,
        weight=weight,
        from_location=location,
        to_location=location if action != 'out' else None,
        operator=current_user.username,
    )

    db.session.add(trans)
    db.session.flush()

    _update_item_status(target_type, target_id, action,
                        to_location=trans.to_location)
    db.session.commit()

    action_labels = {'in': '入库', 'out': '出库', 'lock': '锁定', 'unlock': '解锁'}
    flash(f'{item_name} 已{action_labels.get(action, action)}', 'success')
    return item


@main_bp.route('/raw/<int:id>/action/<action>', methods=['POST'])
@login_required
def raw_action(id, action):
    if action not in ('in', 'out', 'lock', 'unlock'):
        flash('无效操作', 'danger')
        return redirect(url_for('main.raw_detail', id=id))
    _quick_action('raw', id, action)
    return redirect(url_for('main.raw_detail', id=id))


@main_bp.route('/semi/<int:id>/action/<action>', methods=['POST'])
@login_required
def semi_action(id, action):
    if action not in ('in', 'out', 'lock', 'unlock'):
        flash('无效操作', 'danger')
        return redirect(url_for('main.semi_detail', id=id))
    _quick_action('semi', id, action)
    return redirect(url_for('main.semi_detail', id=id))


@main_bp.route('/finished/<int:id>/action/<action>', methods=['POST'])
@login_required
def finished_action(id, action):
    if action not in ('in', 'out', 'lock', 'unlock'):
        flash('无效操作', 'danger')
        return redirect(url_for('main.finished_detail', id=id))
    _quick_action('finished', id, action)
    return redirect(url_for('main.finished_detail', id=id))


# ══════════════════════════════════════════════════════════
# 大系统联动 — 订单浏览
# ══════════════════════════════════════════════════════════

@main_bp.route('/orders')
@login_required
def order_list():
    """浏览大系统订单（只读）"""
    search = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 30

    orders, total = wf_search_orders(search=search, status=status, page=page, per_page=per_page)

    from math import ceil
    pages = ceil(total / per_page) if total > 0 else 1

    pagination = {
        'page': page, 'pages': pages, 'total': total,
        'has_prev': page > 1, 'has_next': page < pages,
        'prev_num': page - 1, 'next_num': page + 1,
    }

    # 查询关联的库存物料数
    for order in orders:
        order['linked_stock_count'] = Transaction.query.filter_by(
            related_order_id=order['custom_order_no'] or str(order['id'])
        ).count()

    return render_template('orders/list.html',
                           orders=orders, pagination=pagination,
                           search=search, status=status)


@main_bp.route('/orders/<int:order_id>')
@login_required
def order_detail(order_id):
    """订单详情 + 关联的库存物料"""
    order = wf_get_order(order_id)
    if not order:
        flash('未找到该订单，可能大系统数据库不可用', 'warning')
        return redirect(url_for('main.order_list'))

    stones = get_order_stones(order_id)

    # 查询关联的库存流水
    order_refs = [order.get('custom_order_no')]
    if str(order_id) not in order_refs:
        order_refs.append(str(order_id))
    order_refs = [r for r in order_refs if r]

    linked_transactions = []
    if order_refs:
        linked_transactions = Transaction.query.filter(
            Transaction.related_order_id.in_(order_refs)
        ).order_by(Transaction.created_at.desc()).all()

        if not current_user.is_admin:
            for t in linked_transactions:
                t.cost_recorded = None

    return render_template('orders/detail.html',
                           order=order, stones=stones,
                           linked_transactions=linked_transactions)


@main_bp.route('/orders/search-api')
@login_required
def order_search_api():
    """JSON API: 搜索订单供下拉选择（用于出入库表单）"""
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return {'results': []}

    orders, _ = wf_search_orders(search=q, per_page=15)
    results = []
    for o in orders:
        results.append({
            'id': o['id'],
            'order_no': o.get('custom_order_no', ''),
            'title': o.get('title', ''),
            'customer': o.get('customer_name', ''),
            'status': wf_status_display(o.get('status', '')),
        })
    return {'results': results}
